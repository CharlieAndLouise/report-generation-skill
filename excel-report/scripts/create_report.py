#!/usr/bin/env python3
"""
Bundled Excel report builder for the excel-report skill.
Handles logo placement, header layout, and data formatting.

Usage:
    python create_report.py \
        --title "Report Title" \
        --summary "Summary sentence" \
        --data /tmp/data.json \
        --logo resources/company_logo.png \
        --output output/report_20250115_120000.xlsx
"""

import argparse
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "D6E4F0"
SUMMARY_BG = "F2F2F2"
SUMMARY_FG = "595959"

CURRENCY_KEYWORDS = {"salary", "pay", "wage", "compensation", "income", "bonus", "rate"}


def _is_currency_col(key: str) -> bool:
    return any(kw in key.lower() for kw in CURRENCY_KEYWORDS)


def create_report(title: str, summary: str, data: list, logo_path: str, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Logo + title area (rows 1-2) ---
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20
    ws.column_dimensions["A"].width = 4   # narrow col for logo left margin
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 5   # spacer between logo and title

    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 100
        img.height = 55
        img.anchor = "A1"
        ws.add_image(img)

    title_cell = ws["D1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color=HEADER_BG)
    title_cell.alignment = Alignment(vertical="center", wrap_text=False)

    # --- Summary row (row 3) ---
    ws.row_dimensions[3].height = 18
    n_cols = len(data[0]) if data else 4
    end_col = get_column_letter(max(n_cols, 4))
    ws.merge_cells(f"A3:{end_col}3")
    summary_cell = ws["A3"]
    summary_cell.value = summary
    summary_cell.font = Font(italic=True, size=10, color=SUMMARY_FG)
    summary_cell.fill = PatternFill("solid", fgColor=SUMMARY_BG)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    if not data:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        print(f"Report saved: {output_path}")
        return

    headers = list(data[0].keys())

    # --- Column headers (row 4) ---
    ws.row_dimensions[4].height = 22
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, size=11)

    for col_idx, key in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = key.replace("_", " ").title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data rows (row 5+) ---
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)

    for row_idx, record in enumerate(data, start=5):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = record.get(key, "")

            if _is_currency_col(key):
                try:
                    cell.value = float(str(value).replace(",", "").replace("$", ""))
                    cell.number_format = '$#,##0'
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value

            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")

    # --- Auto-fit column widths ---
    for col_idx, key in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(key.replace("_", " ").title())
        for row_idx in range(5, len(data) + 5):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Report saved: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate formatted Excel report")
    parser.add_argument("--title", required=True, help="Report title shown beside logo")
    parser.add_argument("--summary", required=True, help="One-sentence summary row")
    parser.add_argument("--data", required=True, help="Path to JSON file (list of dicts)")
    parser.add_argument("--logo", default="resources/company_logo.png", help="Path to logo PNG")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.data, encoding="utf-8") as f:
        rows = json.load(f)

    create_report(args.title, args.summary, rows, args.logo, args.output)
